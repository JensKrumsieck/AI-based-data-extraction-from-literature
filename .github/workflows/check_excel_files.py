original_data = "data/05_manual_tabular"
fake_data = "04_manual_tabular"

import json
import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter


def compare_excel(file1, file2) -> bool:
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)

    sheets1, sheets2 = set(wb1.sheetnames), set(wb2.sheetnames)
    check = True

    for name in sorted(sheets1 | sheets2):
        if name not in sheets1:
            print(f"Sheet only in {file2}: '{name}'")
            check = False
            continue
        if name not in sheets2:
            print(f"Sheet only in {file1}: '{name}'")
            check = False
            continue

        ws1, ws2 = wb1[name], wb2[name]
        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v1 = ws1.cell(r, c).value
                v2 = ws2.cell(r, c).value
                if v1 != v2:
                    col_letter = get_column_letter(c)
                    print(f"[{name}] row={r} col={col_letter} val1={v1!r} val2={v2!r}")
                    check = False
    return check

def compare_jsonl(file1, file2):
    """Compare two JSONL files line by line; print whether they match and any diffs."""
    with open(file1, encoding="utf-8") as f:
        records1 = [json.loads(line) for line in f if line.strip()]
    with open(file2, encoding="utf-8") as f:
        records2 = [json.loads(line) for line in f if line.strip()]

    if records1 == records2:
        print("Files are identical.")
        return

    if len(records1) != len(records2):
        print(f"Line count differs: {len(records1)} vs {len(records2)} (comparing overlap only)")

def find_files(folder, ext):
    files = {}
    for root, _, filenames in os.walk(folder):
        for f in filenames:
            if f.lower().endswith(ext):
                files[f] = os.path.join(root, f)
    return files


def find_and_pair(dir1, dir2, ext):
    orig_files = find_files(dir1, ext)
    fake_files = find_files(dir2, ext)

    common = sorted(set(orig_files) & set(fake_files))
    only_orig = sorted(set(orig_files) - set(fake_files))
    only_fake = sorted(set(fake_files) - set(orig_files))

    for name in only_orig:
        print(f"Only in {original_data}: {name}")
        exit(1)
    for name in only_fake:
        print(f"Only in {fake_data}: {name}")
        exit(1)

    return [(name, orig_files[name], fake_files[name]) for name in common]


if __name__ == "__main__":

    for name, f1, f2 in find_and_pair(original_data, fake_data, ".xlsx"):
        print(f"\n=== Comparing {name} ===")
        if not compare_excel(f1, f2):
            sys.exit(1)
        else: 
            print("passed ✔️")