# -*- coding: utf-8 -*-
"""
Created on Mon May  4 11:52:52 2026

@author: xinxin
"""

import os, sys, json
import re
from pathlib import Path
import pandas as pd
import numpy as np

# from difflib import SequenceMatcher
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
import argparse


def normalize_string(s):
    """Normalize string for comparison: lower case, ..."""
    if not isinstance(s, str):  # if it is not a string, return it as is
        return s
    s = re.sub(
        r"\(.*?\)", "", s
    )  # Remove parentheses and their contents (e.g., " (RCBD)")
    s = re.sub(
        r"[^a-zA-Z0-9\s]", "", s
    ).lower()  # Convert to lowercase and remove special characters
    return " ".join(s.split())  # Strip extra whitespace and return


def extract_numbers(s):
    """Extracts all numbers from a string"""
    if not isinstance(s, str):
        return [str(s)] if pd.notna(s) else []
    return re.findall(r"\d+\.?\d*", s)


# def are_similar(str1, str2, threshold=0.8):
#     """Checks if two strings are similar above a certain percentage"""
#     if not str1 or not str2:
#         return str1 == str2
#     return SequenceMatcher(None, str1, str2).ratio() >= threshold


def get_classification(val_manual, val_model):

    # 1. True Negative: both are empty
    if pd.isna(val_manual) and pd.isna(val_model):
        return "TN"

    # 2. False Positive: manual is empty, model has text
    if pd.isna(val_manual) and not pd.isna(val_model):
        return "FP"

    # 3. False Negative 2: model is empty, manual has text -> model missed content
    if pd.isna(val_model) and not pd.isna(val_manual):
        return "FN2"

    numbers_from_manual = extract_numbers(val_manual)
    numbers_from_model = extract_numbers(val_model)

    # If there are numbers in both, they must match exactly
    try:
        float_numbers_manual = [float(num) for num in numbers_from_manual]
        float_numbers_model = [float(num) for num in numbers_from_model]
        if float_numbers_manual != float_numbers_model:
            return "FN1"
    except (ValueError, TypeError):
        if numbers_from_manual != numbers_from_model:
            return "FN1"

    # If numbers match (or there are no numbers), proceed with text fuzzy matching
    norm_manual = normalize_string(val_manual)
    norm_model = normalize_string(val_model)

    if norm_manual == norm_model:
        return "TP"
    else:
        return "TO CHECK!"


def compare_dataframes(man_df, mod_df, filename):
    """Compare two dataframes and return a list of rows with classifications for each column."""
    results = []

    # Use all columns from the manual dataframe
    manual_cols = man_df.columns.tolist()
    num_rows = max(len(man_df), len(mod_df))

    for index in range(num_rows):
        # Start the row with the filename
        row_result = {"Manual File": filename}

        for col in manual_cols:
            val_manual = man_df.iloc[index][col] if index < len(man_df) else np.nan

            # Check if column exists in model dataframe
            if col in mod_df.columns:
                val_model = mod_df.iloc[index][col] if index < len(mod_df) else np.nan
                classification = get_classification(val_manual, val_model)
            else:
                # Column missing in model entirely
                classification = "FN2" if pd.notna(val_manual) else "TN"

            row_result[col] = classification

        results.append(row_result)

    return results


parser = argparse.ArgumentParser("evaluate model")
parser.add_argument("-r", "--reference_data", help="manual tabular", required=True)
parser.add_argument(
    "-o",
    "--output_folder",
    help="Folder for all intermediate + final output",
    required=True,
)

# model keys to automatically generate arguments
model_keys = [
    "context_metadata",
    "fields",
    "genotypes",
    "plantings",
    "irrigations",
    "fertilizers",
    "harvests",
    "plot_details",
]

for item in model_keys:
    # arguments are not required
    parser.add_argument(
        f"--{item}_dir",
        help=f"Folder for ICASA {item} category",
    )

args = parser.parse_args()
manual_path = Path(args.reference_data)
output_dir = args.output_folder
model_genotypes_path = args.genotypes_dir
model_context_metadata_path = args.context_metadata_dir
model_fields_path = args.fields_dir
model_plantings_path = args.plantings_dir
model_irrigations_path = args.irrigations_dir
model_fertilizer_path = args.fertilizers_dir
model_plot_path = args.plot_details_dir
model_harvest_path = args.harvests_dir

# uncomment to overwrite args
# reference_path = Path(r"C:\your\path\to")
#
# model_path = reference_path / "data"
# manual_path = reference_path / "05_manual_tabular"
# model_genotypes_path = reference_path / "08_llm_output_tabular/genotypes"
# model_context_metadata_path = reference_path / "08_llm_output_tabular/context_metadata"
# model_fields_path = reference_path / "08_llm_output_tabular/fields"
# model_plantings_path = reference_path / "08_llm_output_tabular/plantings"
# model_irrigations_path = reference_path / "08_llm_output_tabular/irrigations"
# model_fertilizer_path = reference_path / "08_llm_output_tabular/fertilizeration"
# model_plot_path = reference_path / "08_llm_output_tabular/plot_details"
# model_harvest_path = reference_path / "08_llm_output_tabular/harvests"
# output_dir = os.path.join(reference_path, "..", "output")


os.makedirs(output_dir, exist_ok=True)

CATEGORIES = {
    "genotypes":        ("GENOTYPES_comparison.xlsx",   "GENOTYPES",       model_genotypes_path),
    "context_metadata": ("METADATA_comparison.xlsx",    "EXP_METADATA", model_context_metadata_path),
    "fields":           ("FIELDS_comparison.xlsx",      "FIELDS",          model_fields_path),
    "plantings":        ("PLANTINGS_comparison.xlsx",   "PLANTINGS",       model_plantings_path),
    "irrigations":      ("IRRIGATIONS_comparison.xlsx", "IRRIGATIONS",     model_irrigations_path),
    "fertilizers":      ("FERTILIZERS_comparison.xlsx", "FERTILIZERS",     model_fertilizer_path),
    "plot_details":     ("PLOT_DETAILS_comparison.xlsx","PLOT_DETAILS",     model_plot_path),
    "harvests":         ("HARVESTS_comparison.xlsx",    "HARVESTS",        model_harvest_path),
}

for category, (out_filename, sheet_to_read, model_folder_raw) in CATEGORIES.items():
    if model_folder_raw is None:
        print(f"Skipping {category}: no folder provided.")
        continue
    output_file = os.path.join(output_dir, out_filename)
    model_folder = Path(model_folder_raw)
    
    print(f"\n=== Processing {category} ===")
    all_results = []

    # Create a mapping for model files by prefix for easier lookup
    model_files = {
        f.split("_")[0]: f for f in os.listdir(model_folder) if f.endswith(".xlsx")
    }
    manual_files = [f for f in os.listdir(manual_path) if f.endswith(".xlsx")]

    header_written = False
    manual_df = None
    model_df = None

    for f in manual_files:
        prefix = f.split(".")[0]
        if prefix in model_files:
            print(f"Comparing {prefix}...")
            model_filename = model_files[prefix]

            try:
                manual_df = pd.read_excel(manual_path / f, sheet_name=sheet_to_read)
                model_df = pd.read_excel(model_folder / model_filename)

                if manual_df is not None and model_df is not None:
                    file_diffs = compare_dataframes(manual_df, model_df, f)
                    if file_diffs:
                        all_results.extend(file_diffs)

            except Exception as e:
                print(f"Error processing {prefix}: {e}")

        if all_results:
            final_df = pd.DataFrame(all_results)

            # Reorder columns to put "Manual File" first
            cols = ["Manual File"] + [c for c in final_df.columns if c != "Manual File"]
            final_df = final_df[cols]

            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                final_df.to_excel(writer, index=False, sheet_name="Results")
                worksheet = writer.sheets["Results"]

                # Define styles
                red_f = Font(color="8b0000")
                yellow_bg = PatternFill(
                    start_color="fff000", end_color="fff000", fill_type="solid"
                )
                orange_bg = PatternFill(
                    start_color="ff8f00", end_color="ff8f00", fill_type="solid"
                )
                green_bg = PatternFill(
                    start_color="90ee90", end_color="90ee90", fill_type="solid"
                )
                blue_bg = PatternFill(
                    start_color="87ceeb", end_color="87ceeb", fill_type="solid"
                )
                red_bg = PatternFill(
                    start_color="f08080", end_color="f08080", fill_type="solid"
                )

                # Apply rule: Match "TO CHECK" (must match the string in get_classification)
                worksheet.conditional_formatting.add(
                    f"A1:Z{len(final_df) + 1}",
                    CellIsRule(operator="equal", formula=['"TO CHECK!"'], font=red_f),
                )
                # Yellow background for "FN1"
                worksheet.conditional_formatting.add(
                    f"A1:Z{len(final_df) + 1}",
                    CellIsRule(operator="equal", formula=['"FN1"'], fill=yellow_bg),
                )

                # Orange background for "FN2"
                worksheet.conditional_formatting.add(
                    f"A1:Z{len(final_df) + 1}",
                    CellIsRule(operator="equal", formula=['"FN2"'], fill=orange_bg),
                )

                # RED background for "FP"
                worksheet.conditional_formatting.add(
                    f"A1:Z{len(final_df) + 1}",
                    CellIsRule(operator="equal", formula=['"FP"'], fill=red_bg),
                )

                # BLUE background for "TN"
                worksheet.conditional_formatting.add(
                    f"A1:Z{len(final_df) + 1}",
                    CellIsRule(operator="equal", formula=['"TN"'], fill=blue_bg),
                )

                # GREEN background for "TP"
                worksheet.conditional_formatting.add(
                    f"A1:Z{len(final_df) + 1}",
                    CellIsRule(operator="equal", formula=['"TP"'], fill=green_bg),
                )

            print(f"Processing finished. Results saved to {output_file}")
        else:
            print("No results to save.")


print("Processing finished.")
